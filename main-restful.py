import string

import requests
import json
from openpyxl import load_workbook
from tkinter import Tk
from tkinter.filedialog import askopenfilename
import time
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

usernamePass = 'username=prtgadmin&passhash=2446819464'
global host
host = '192.168.16.89'

def checkStatus(id):
    strid = str(id)
    if strid == '401':
        print('401 Unauthorized Access')
    elif strid == '200':
        print('200 Success')
    elif strid == '302':
        print('302 Found')
    elif strid == '404':
        print('404 Not Found')
    else:
        print('Unknown Error')

def add_device(id_clone, name, host_ip, id_group):
    response = requests.post('https://'+host+'/api/duplicateobject.htm?id='+id_clone+'&name='+name+'&host='+host_ip+'&targetid='+str(id_group)+'&'+usernamePass, verify=False)
    checkStatus(response.status_code)

def addAlldevice_to_specificGroup(wb):
    wb = loadExcel(wb)
    for i in range(4, wb.max_row + 1):
        cell_obj = wb.cell(row=i, column=3)
        if (checkDuplicateDevice(cell_obj.value) == True):
            print("Device exists. SKIPPED!!!")
        elif (checkDuplicateDevice(cell_obj.value) == False ):
            add_device('2929', cell_obj.value, wb.cell(row=i, column=4).value, getGroupID(wb.cell(row=i, column=2).value))

"""    for i in range(1, wb.max_row + 1):
        cell_obj = wb.cell(row=i, column=1)
        if (checkDuplicateDevice(wb.cell(row=i, column=2).value) == True):
            continue
        elif (checkDuplicateDevice(wb.cell(row=i, column=2).value) == False):
            add_device('2929', wb.cell(row=i, column=2).value, wb.cell(row=i, column=3).value, getGroupID(cell_obj.value))
"""

def addGroup(id_clone, name, destination_id_group):
    response = requests.post('https://'+host+'/api/duplicateobject.htm?id='+id_clone+'&name='+stringProcessor(name)+'&targetid='+str(destination_id_group)+'&'+usernamePass, verify=False)
    checkStatus(response.status_code)


def getDeviceList():
    response = requests.get('https://'+host+'/api/table.xml?content=sensors&columns=objid,group,device,status&count=1000&'+usernamePass, verify=False)
    print(response.text)

def getAllGroupID():
    response = requests.get('https://'+host+'/api/table.json?content=groups&output=json&columns=objid,group&noraw=1&usecaption=true&count=1000&'+usernamePass, verify=False)
    print(response.text)

def stringProcessor(string):
    newstring = string.replace(" ", "+")
    newstring = newstring.replace("&", "%26")
    return newstring

def getGroupID(group_name):
    print("Getting "+ group_name +" ID")
    response = requests.get('https://' + host + '/api/table.json?content=groups&output=json&columns=objid,name&count=1000&filter_name='+stringProcessor(group_name)+'&'+usernamePass, verify=False)
    groups = json.loads(response.text)
    print("ID is "+ str(groups['groups'][0]['objid']))
    return groups['groups'][0]['objid']

def addMultipleGroupID(sheet):
    wb = loadExcel(sheet)
    for i in range(4, wb.max_row + 1):
        cell_obj = wb.cell(row=i, column=1)
        if (checkDuplicateGroup(cell_obj.value) == True):                                   # Check group exists
            cell_obj2 = wb.cell(row=i, column=2)                                            # Retrieve second column cell
            if (checkDuplicateGroup(cell_obj2.value) == True):                              # If second column cell values exists
                print("Sub group exists. SKIPPED!!!")
                continue
            else:                                                                           # Value not exist. Proceed to add sub group.
                addGroup('2213', cell_obj2.value, getGroupID(cell_obj.value))
        elif (checkDuplicateGroup(cell_obj.value) == False ):                               # Group not exists
            addGroup('2213', cell_obj.value, '1')                                           # Add group to main
            cell_obj2 = wb.cell(row=i, column=2)                                            # Retrieve second column cell
            addGroup('2213', cell_obj2.value, getGroupID(cell_obj.value))


def matchwithGroupName(group_name):
    response = requests.get('https://'+host+'/api/table.json?content=groups&output=json&columns=objid,group&noraw=1&usecaption=true&count=1000&'+usernamePass, verify=False)
    #print(response.text)
    groups = json.loads(response.text)
    print(groups['groups'][1]['objid'])
    max_group_length = (len(groups['groups']))
    for j in max_group_length:
        if (group_name == groups['groups'][j]['group']):
            print("Group name found. Returning object ID to add device")
            return groups['groups'][j]['objid']
        else:
            print("Group name not found.")

def resumeObject(id):
    response = requests.post('https://'+host+'/api/pause.htm?id='+str(id)+'&action=1&count=1000&'+usernamePass, verify=False)
    print(response.text)

def resumeMultipleObject():
    response_get = requests.get('https://'+host+'/api/table.json?content=groups&output=json&columns=objid,group&noraw=1&usecaption=true&count=1000&'+usernamePass, verify=False)
    groups_group = json.loads(response_get.text)
    for i in len(groups_group):
        if(groups_group['groups']['i']['objid'] != ''):
            resumeObject(groups_group['groups'][i]['objid'])

def getFilePath():
    filepath = input("Enter the path by browsing file (Y) or manual entry (N)? ")
    while(1):
        if (filepath == 'Y' or filepath == 'y'):
            filepaths = askopenfilename()
            print(filepaths)
            break
        elif (filepath == 'N' or filepath == 'n'):
            filepaths = input("Enter path: ")
        else:
            print("Invalid input")
            break
    return filepaths

def loadExcel(filepath):
    wb = load_workbook(filepath)
    sheet = wb.active
    return sheet

def checkDuplicateGroup(name):
    response = requests.get('https://' + host + '/api/table.json?content=groups&output=json&columns=objid,group&noraw=1&usecaption=true&count=1000&'+usernamePass, verify=False)
    groups = json.loads(response.text)
    max_group_lengths = (len(groups['groups']))
    print(max_group_lengths)
    for i in range(max_group_lengths):
        if (name == groups['groups'][i]['group']):
            print("Group name exists. SKIPPED!!! Proceed adding sub group")
            return True
        elif (max_group_lengths == 0):
            return None
    return False


def checkDuplicateDevice(name):
    response = requests.get('https://' + host + '/api/table.json?content=sensors&output=json&columns=objid,device&noraw=1&usecaption=true&count=1000&'+usernamePass, verify=False)
    groups = json.loads(response.text)
    print(response.text)
    max_group_lengths = (len(groups['sensors']))
    print(max_group_lengths)
    for i in range(max_group_lengths):
        if (name == groups['sensors'][i]['device']):
            print("Device "+name+" name exists. SKIPPED!!!")
            return True
        elif (max_group_lengths == 0):
            return None
    return False

def checkStringPopulate(sheet):
    wb = loadExcel(sheet)
    d = dict()
    sum = 0
    for i in range(4, wb.max_row + 1):
        cell_obj = wb.cell(row=i, column=23)
        strings = cell_obj.value

        #for word in strings:
        if strings in d:
            d[strings] = d[strings] + 1
        else:
            d[strings] = 1
    for key in list(d.keys()):
        print(key, ":", d[key])
        sum += d[key]
    print("Total devices: " + str(sum))

def setHost(address):
    host = address

def main():
    root = Tk()
    while (1):
        print("PRTG Automated Deployment Application")
        select_action = input("Select actions: \n (1) Add group \n (2) Add devices \n (3) Populate devices from Excel file only \n (4) Exit \n")
        if (select_action == '1'):
            filepath = getFilePath()
            addMultipleGroupID(filepath)
            print("Add groups ended")
        elif (select_action == '2'):
            filepath = getFilePath()
            addAlldevice_to_specificGroup(filepath)
            print("Add devices ended")
        elif (select_action == '3'):
            filepath = getFilePath()
            checkStringPopulate(filepath)
        elif (select_action == '4'):
            break
        else:
            print("Invalid input")
    root.withdraw()
    root.destroy()
    print("Operation ended. You may close the application")
    time.sleep(1000)
if __name__ == "__main__":
    main()



